from rdflib import Graph, Namespace, URIRef, BNode, Literal
from elasticsearch import Elasticsearch
import os
import requests
from urllib.parse import urljoin, urlparse, unquote
from pdfminer.high_level import extract_text as pdf_extract_text
from docx import Document
from openpyxl import load_workbook
import io
import logging
import time
from functools import lru_cache, wraps
import dotenv
from pathlib import Path
import argparse

# Load environment variables from .env file
dotenv.load_dotenv()

# Define namespaces
LDTO = Namespace("https://data.razu.nl/def/ldto/")
SKOS = Namespace("http://www.w3.org/2004/02/skos/core#")
DCT = Namespace("http://purl.org/dc/terms/")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_last_segment(uri):
    """Extract the last segment of a URI."""
    return str(uri).split('/')[-1]

def normalize_label(label):
    """Normalize a label to be used as a field name"""
    return label.lower().replace(' ', '_')

def dereference_uri(uri):
    """Dereference a URI and return a Graph with its content"""
    try:
        response = requests.get(uri, headers={'Accept': 'text/turtle'})
        if response.status_code == 429:  # Rate limit hit
            time.sleep(1)  # Wait a bit
            response = requests.get(uri, headers={'Accept': 'text/turtle'})
            
        if response.status_code != 200:
            logger.warning(f"Failed to dereference {uri}: {response.status_code}")
            return None
            
        # Parse the response into a temporary graph
        temp_g = Graph()
        temp_g.parse(data=response.text, format='turtle')
        return temp_g
        
    except Exception as e:
        logger.error(f"Error dereferencing {uri}: {str(e)}")
        return None

def get_skos_triple_value(g, subject, predicate):
    """Get a single value from a SKOS triple"""
    for _, _, value in g.triples((URIRef(subject), predicate, None)):
        return str(value)
    return None

@lru_cache(maxsize=1000)
def get_skos_label(uri):
    """Fetch SKOS prefLabel for a given URI. Results are cached using lru_cache."""
    g = dereference_uri(uri)
    if not g:
        return None
        
    label = get_skos_triple_value(g, uri, SKOS.prefLabel)
    if not label:
        logger.warning(f"No prefLabel found for {uri}")
    return label

@lru_cache(maxsize=1000)
def get_beperking_gebruik_labels(uri):
    """Get the scheme and label for a beperkingGebruik URI by dereferencing the URIs"""
    logger.info(f"Dereferencing URI: {uri}")
    
    # Get the beperking type label and scheme
    g = dereference_uri(uri)
    if not g:
        return None
    
    # Get the prefLabel and inScheme
    type_label = get_skos_triple_value(g, uri, SKOS.prefLabel)
    scheme_uri = get_skos_triple_value(g, uri, SKOS.inScheme)
    
    if not type_label or not scheme_uri:
        logger.warning(f"No type_label or scheme found for {uri}")
        return None
    
    # Now dereference the scheme to get its label
    logger.info(f"Dereferencing scheme: {scheme_uri}")
    scheme_g = dereference_uri(scheme_uri)
    if not scheme_g:
        return None
    
    # Get the scheme label
    scheme_label = get_skos_triple_value(scheme_g, scheme_uri, SKOS.prefLabel)
    if not scheme_label:
        logger.warning(f"No scheme label found for {scheme_uri}")
        return None
    
    normalized = normalize_label(scheme_label)
    logger.info(f"Found labels - scheme: {scheme_label} -> {normalized}, type: {type_label}")
    
    return {
        'scheme': scheme_uri,
        'scheme_label': normalized,
        'type_label': type_label
    }

def get_scheme_labels(g):
    """Get labels for all schemes and their types."""
    query = """
    PREFIX ldto: <https://data.razu.nl/def/ldto/>
    PREFIX skos: <http://www.w3.org/2004/02/skos/core#>
    SELECT ?scheme ?scheme_label ?type ?type_label
    WHERE {
        ?scheme a skos:ConceptScheme ;
                skos:prefLabel ?scheme_label .
        ?type skos:inScheme ?scheme ;
              skos:prefLabel ?type_label .
    }
    """
    scheme_labels = {}
    for row in g.query(query):
        scheme_uri = str(row['scheme'])
        scheme_label = str(row['scheme_label'])
        type_label = str(row['type_label'])
        
        # Convert scheme label to snake_case for use as field name
        field_name = scheme_label.lower().replace(' ', '_')
        scheme_labels[scheme_uri] = field_name
        
        logger.info(f"Found labels - scheme: {scheme_label} -> {field_name}, type: {type_label}")
    
    return scheme_labels

def get_dekking_types(g):
    """Get all dekkingInTijdType values."""
    query = """
    PREFIX ldto: <https://data.razu.nl/def/ldto/>
    PREFIX skos: <http://www.w3.org/2004/02/skos/core#>
    SELECT ?type ?label
    WHERE {
        ?type a ldto:DekkingInTijdType ;
              skos:prefLabel ?label .
    }
    """
    dekking_types = {}
    for row in g.query(query):
        type_uri = str(row['type'])
        label = str(row['label'])
        dekking_types[type_uri] = label
        logger.info(f"Found dekkingInTijdType: {label} ({type_uri})")
    return dekking_types

def find_root_archive(g, doc_uri, documents):
    """Find the root archive by following ldto:isOnderdeelVan until we reach the top level."""
    current_uri = doc_uri
    while True:
        # Get parent URI through ldto:isOnderdeelVan
        parent_query = """
        PREFIX ldto: <https://data.razu.nl/def/ldto/>
        SELECT ?parent
        WHERE {
            <%s> ldto:isOnderdeelVan ?parent .
        }
        """ % current_uri
        
        results = g.query(parent_query)
        parent_uris = [str(row['parent']) for row in results]
        
        if not parent_uris:  # No parent found, this is the root
            # Get the naam of the current document
            naam_query = """
            PREFIX ldto: <https://data.razu.nl/def/ldto/>
            SELECT ?naam
            WHERE {
                <%s> ldto:naam ?naam .
            }
            """ % current_uri
            naam_results = g.query(naam_query)
            for row in naam_results:
                return str(row['naam'])
            return None  # No naam found
            
        current_uri = parent_uris[0]  # Follow the first parent

def extract_text_from_pdf(file_path):
    """Extract text from a PDF file"""
    try:
        return pdf_extract_text(file_path)
    except Exception as e:
        logger.error(f"Error extracting text from PDF {file_path}: {e}")
        return None

def extract_text_from_docx(file_path):
    """Extract text from a DOCX file"""
    try:
        doc = Document(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])
    except Exception as e:
        logger.error(f"Error extracting text from DOCX {file_path}: {e}")
        return None

def extract_text_from_xlsx(file_path):
    """Extract text from an XLSX file"""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        texts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.rows:
                texts.extend(str(cell.value) for cell in row if cell.value)
        return "\n".join(texts)
    except Exception as e:
        logger.error(f"Error extracting text from XLSX {file_path}: {e}")
        return None

def get_file_content(url):
    """Get text content from a file in the docs directory"""
    try:
        # Extract filename from URL and decode it
        filename = unquote(os.path.basename(urlparse(url).path))
        file_path = os.path.join('docs', filename)
        
        if not os.path.exists(file_path):
            logger.warning(f"File not found: {file_path}")
            return None
            
        # Determine file type and extract text
        ext = os.path.splitext(filename)[1].lower()
        if ext == '.pdf':
            return extract_text_from_pdf(file_path)
        elif ext == '.docx':
            return extract_text_from_docx(file_path)
        elif ext == '.xlsx':
            return extract_text_from_xlsx(file_path)
        else:
            logger.warning(f"Unsupported file type: {ext}")
            return None
            
    except Exception as e:
        logger.error(f"Error processing file from URL {url}: {e}")
        return None

def create_index(es, index_name, dekking_types, scheme_labels):
    """Create Elasticsearch index with mapping for LDTO fields if it doesn't exist"""
    # Check if index exists
    if es.indices.exists(index=index_name):
        logger.info(f"Index {index_name} already exists")
        return
        
    # Create index with mapping
    mapping = {
        "_source": {
            "excludes": ["full_text"]
        },
        "properties": {
            "id": {"type": "keyword"},
            "naam": {
                "type": "text",
                "analyzer": "dutch",
                "fields": {
                    "keyword": {"type": "keyword"}
                }
            },
            "omschrijving": {
                "type": "text",
                "analyzer": "dutch"
            },
            "classificatie": {
                "type": "text",
                "analyzer": "dutch",
                "fields": {
                    "keyword": {"type": "keyword"}
                }
            },
            "archiefvormer": {
                "type": "text",
                "analyzer": "dutch",
                "fields": {
                    "keyword": {"type": "keyword"}
                }
            },
            "aggregatieniveau": {
                "type": "text",
                "analyzer": "dutch",
                "fields": {
                    "keyword": {"type": "keyword"}
                }
            },
            "archief": {
                "type": "text",
                "analyzer": "dutch",
                "fields": {
                    "keyword": {"type": "keyword"}
                }
            },
            "full_text": {
                "type": "text",
                "analyzer": "dutch",
                "store": False,
                "index_options": "docs",
                "norms": True
            },
            "bestand_url": {"type": "keyword"},
            # Metadata fields - for search and faceting
            "classificatie_uri": {
                "type": "keyword",
                "index": False
            },
            "archiefvormer_uri": {
                "type": "keyword",
                "index": False
            },
            "aggregatieniveau_uri": {
                "type": "keyword",
                "index": False
            },
            # Hierarchical relationships - not searchable
            "bevat_onderdeel": {
                "type": "keyword",
                "index": False
            },
            "is_onderdeel_van": {
                "type": "keyword",
                "index": False
            }
        }
    }
    
    # Add date range fields for each dekking type
    for dekking_type, field_name in dekking_types.items():
        mapping["properties"][f"{field_name}_range"] = {
            "type": "date_range",
            "format": "yyyy-MM-dd||yyyy"
        }
        
    # Add fields for each scheme
    for scheme_uri, scheme_label in scheme_labels.items():
        mapping["properties"][scheme_label] = {
            "type": "keyword"
        }
        
    # Create index with mapping
    es.indices.create(
        index=index_name,
        mappings=mapping,
        settings={
            "analysis": {
                "analyzer": {
                    "dutch": {
                        "tokenizer": "standard",
                        "filter": ["lowercase", "dutch_stop", "dutch_stemmer"]
                    }
                },
                "filter": {
                    "dutch_stop": {
                        "type": "stop",
                        "stopwords": "_dutch_"
                    },
                    "dutch_stemmer": {
                        "type": "stemmer",
                        "language": "dutch"
                    }
                }
            }
        }
    )
    logger.info(f"Created index {index_name} with facet mapping")

def create_elasticsearch_client():
    """Create and configure Elasticsearch client"""
    es_host = os.getenv('ES_HOST', 'es.digitopia.nl')
    es_port = int(os.getenv('ES_PORT', '443'))
    es_username = os.getenv('ES_USERNAME')
    es_password = os.getenv('ES_PASSWORD')
    
    if not all([es_username, es_password]):
        raise ValueError("ES_USERNAME and ES_PASSWORD must be set in .env file")
    
    return Elasticsearch(
        f"https://{es_host}:{es_port}",
        basic_auth=(es_username, es_password),
        verify_certs=False
    )

def convert_ttl_to_es(ttl_file):
    """Convert TTL file to Elasticsearch documents."""
    g = Graph()
    g.parse(ttl_file, format="turtle")
    
    # First get all scheme labels
    scheme_labels = get_scheme_labels(g)
    
    # Get all dekking types
    dekking_types = get_dekking_types(g)
    
    # Find all ldto:Informatieobject instances
    query = """
    PREFIX ldto: <https://data.razu.nl/def/ldto/>
    PREFIX skos: <http://www.w3.org/2004/02/skos/core#>
    SELECT DISTINCT ?obj ?naam ?omschrijving ?classificatie ?archiefvormer ?aggregatieniveau 
           ?dekking ?dekkingType ?begin ?eind
           ?bestand ?bestandURL ?bestandNaam ?beperkingen
           ?bevatOnderdeel ?isOnderdeelVan
           ?betrokkene ?betrokkeneActor
           (GROUP_CONCAT(?trefwoord; separator=',') as ?trefwoorden)
    WHERE {
        ?obj a ldto:Informatieobject ;
             ldto:naam ?naam .
        OPTIONAL { ?obj ldto:classificatie ?classificatie }
        OPTIONAL { ?obj ldto:archiefvormer ?archiefvormer }
        OPTIONAL { ?obj ldto:aggregatieniveau ?aggregatieniveau }
        OPTIONAL { ?obj ldto:omschrijving ?omschrijving }
        OPTIONAL { 
            ?obj ldto:dekkingInTijd ?dekking .
            ?dekking ldto:dekkingInTijdType ?dekkingType .
            OPTIONAL { ?dekking ldto:begin ?begin }
            OPTIONAL { ?dekking ldto:eind ?eind }
        }
        OPTIONAL { 
            ?obj ldto:heeftRepresentatie ?bestand .
            ?bestand ldto:URLBestand ?bestandURL
        }
        OPTIONAL { ?obj ldto:beperkingGebruik ?beperkingen }
        OPTIONAL { ?obj ldto:bevatOnderdeel ?bevatOnderdeel }
        OPTIONAL { ?obj ldto:isOnderdeelVan ?isOnderdeelVan }
        OPTIONAL { ?obj ldto:trefwoord ?trefwoord }
        OPTIONAL { 
            ?obj ldto:betrokkene ?betrokkene .
            ?betrokkene ldto:betrokkeneActor ?betrokkeneActor
        }
    }
    GROUP BY ?obj ?naam ?omschrijving ?classificatie ?archiefvormer ?aggregatieniveau 
             ?dekking ?dekkingType ?begin ?eind
             ?bestand ?bestandURL ?bestandNaam ?beperkingen
             ?bevatOnderdeel ?isOnderdeelVan
             ?betrokkene ?betrokkeneActor
    """
    
    results = g.query(query)
    
    # Process results into documents
    documents = {}
    
    for row in results:
        doc_id = extract_last_segment(row['obj'])
        doc = {
            'id': doc_id,
            'naam': str(row['naam']),
            'full_text': str(row['naam'])  # Initialize full_text with naam
        }
        
        # Find and add the root archive name
        root_archive = find_root_archive(g, str(row['obj']), documents)
        if root_archive:
            doc['archief'] = root_archive
        
        # Add optional fields if present
        if row['omschrijving']:
            doc['omschrijving'] = str(row['omschrijving'])
            doc['full_text'] += "\n" + str(row['omschrijving'])
            
        if row['classificatie']:
            classificatie_label = get_skos_label(row['classificatie'])
            if classificatie_label:
                doc['classificatie'] = classificatie_label
                doc['classificatie_uri'] = str(row['classificatie'])
            
        if row['archiefvormer']:
            archiefvormer_label = get_skos_label(row['archiefvormer'])
            if archiefvormer_label:
                doc['archiefvormer'] = archiefvormer_label
                doc['archiefvormer_uri'] = str(row['archiefvormer'])
            
        if row['aggregatieniveau']:
            aggregatieniveau_label = get_skos_label(row['aggregatieniveau'])
            if aggregatieniveau_label:
                doc['aggregatieniveau'] = aggregatieniveau_label
                doc['aggregatieniveau_uri'] = str(row['aggregatieniveau'])
        
        # Process bestand if present
        if row['bestand'] and row['bestandURL']:
            doc['bestand_url'] = str(row['bestandURL'])
            # Extract text from the file and add to full_text
            file_content = get_file_content(str(row['bestandURL']))
            if file_content:
                doc['full_text'] += "\n" + file_content
            
        # Add hierarchical relationships if present
        if row['bevatOnderdeel']:
            doc['bevat_onderdeel'] = extract_last_segment(row['bevatOnderdeel'])
            
        if row['isOnderdeelVan']:
            doc['is_onderdeel_van'] = extract_last_segment(row['isOnderdeelVan'])

        # Add keywords if present
        if row['trefwoorden']:
            doc['trefwoorden'] = [kw.strip() for kw in str(row['trefwoorden']).split(',')]
            doc['full_text'] += "\n" + "\n".join(doc['trefwoorden'])

        # Add betrokkeneActor label if present
        if row['betrokkeneActor'] and not isinstance(row['betrokkeneActor'], (BNode, Literal)):
            actor_label = get_skos_label(str(row['betrokkeneActor']))
            if actor_label:
                logger.info(f"Found betrokkeneActor: {actor_label}")
                doc['full_text'] += "\n" + actor_label

        # Store document
        documents[doc_id] = doc
        
    # Process beperkingGebruik for each document
    for doc_id, doc in documents.items():
        # Get all beperkingGebruik values for this document
        query = f"""
        PREFIX ldto: <https://data.razu.nl/def/ldto/>
        SELECT ?beperking
        WHERE {{
            ?obj ldto:beperkingGebruik ?beperking .
            FILTER(STRENDS(str(?obj), "{doc_id}"))
        }}
        """
        
        beperkingen = []
        for row in g.query(query, initNs={'ldto': LDTO}):
            beperkingen.append(str(row[0]))
            
        if beperkingen:
            logger.info(f"Processing beperkingen for {doc_id}: {','.join(beperkingen)}")
            
            # Process each beperking
            for beperking_uri in beperkingen:
                result = get_beperking_gebruik_labels(beperking_uri)
                if result:
                    # Add the type label to the corresponding scheme field
                    scheme_label = result['scheme_label']
                    if scheme_label not in doc:
                        doc[scheme_label] = []
                    doc[scheme_label].append(result['type_label'])
                    logger.info(f"Added beperking to {scheme_label}: {result['type_label']}")
                    
    # Process dekkingInTijd for each document
    for doc_id, doc in documents.items():
        # Get all dekkingInTijd values for this document
        query = f"""
        PREFIX ldto: <https://data.razu.nl/def/ldto/>
        SELECT ?type ?begin ?eind
        WHERE {{
            ?obj ldto:dekkingInTijd ?dekking .
            ?dekking ldto:dekkingInTijdType ?type .
            OPTIONAL {{ ?dekking ldto:begin ?begin }}
            OPTIONAL {{ ?dekking ldto:eind ?eind }}
            FILTER(STRENDS(str(?obj), "{doc_id}"))
        }}
        """
        
        for row in g.query(query, initNs={'ldto': LDTO}):
            dekking_type, begin, eind = row
            
            if str(dekking_type) in dekking_types:
                field_name = dekking_types[str(dekking_type)]
                
                # Create date range
                date_range = {}
                if begin:
                    date_range['gte'] = str(begin)
                if eind:
                    date_range['lte'] = str(eind)
                    
                if date_range:
                    doc[f"{field_name}_range"] = date_range
                    logger.info(f"Added date range for {doc_id}: {begin} - {eind}")
                    
        logger.info(f"Processed document {doc_id} with fields: {list(doc.keys())}")
        
    return documents, dekking_types, scheme_labels

def index_documents(documents, index_name='ldto-objects'):
    """Index documents in Elasticsearch"""
    es = create_elasticsearch_client()
    
    # Create index
    dekking_types = get_dekking_types(Graph())
    scheme_labels = get_scheme_labels(Graph())
    create_index(es, index_name, dekking_types, scheme_labels)
    
    # Index each document
    for doc in documents.values():
        try:
            es.index(index=index_name, id=doc['id'], document=doc)
            logger.info(f"Indexed document {doc['id']}")
        except Exception as e:
            logger.error(f"Error indexing document {doc['id']}: {str(e)}")
            
    # Refresh index to make documents searchable
    es.indices.refresh(index=index_name)
    logger.info("Refreshed index")

def search_and_print_results(es, index_name, query):
    result = es.search(index=index_name, body={
        "query": {
            "multi_match": {
                "query": query,
                "fields": ["naam", "omschrijving", "classificatie", "archiefvormer", "aggregatieniveau", "full_text"],
                "type": "best_fields"
            }
        }
    })
    print(f"\nSearch for '{query}':")
    for hit in result['hits']['hits']:
        source = hit['_source']
        print(f"- {source['naam']} (Score: {hit['_score']})")
        if 'omschrijving' in source:
            print(f"  Omschrijving: {source['omschrijving']}")
        if 'classificatie' in source:
            print(f"  Classificatie: {source['classificatie']}")
        if 'aggregatieniveau' in source:
            print(f"  Aggregatieniveau: {source['aggregatieniveau']}")
        if 'archiefvormer' in source:
            print(f"  Archiefvormer: {source['archiefvormer']}")
        if 'archief' in source:
            print(f"  Archief: {source['archief']}")

if __name__ == "__main__":
    import argparse
    
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Convert TTL file to Elasticsearch index')
    parser.add_argument('ttl_file', help='Path to TTL file')
    args = parser.parse_args()
    
    # Convert TTL to Elasticsearch documents
    documents, dekking_types, scheme_labels = convert_ttl_to_es(args.ttl_file)
    
    # Print found documents
    print(f"\nFound {len(documents)} Informatieobject instances:\n")
    for doc in documents.values():
        print(f"ID: {doc['id']}")
        print(f"Naam: {doc['naam']}")
        if 'omschrijving' in doc:
            print(f"Omschrijving: {doc['omschrijving']}")
        if 'classificatie' in doc:
            print(f"Classificatie: {doc['classificatie']}")
        if 'aggregatieniveau' in doc:
            print(f"Aggregatieniveau: {doc['aggregatieniveau']}")
        if 'archiefvormer' in doc:
            print(f"Archiefvormer: {doc['archiefvormer']}")
        if 'archief' in doc:
            print(f"Archief: {doc['archief']}")
        print()
    
    # Index documents
    index_documents(documents, 'ldto-objects')
